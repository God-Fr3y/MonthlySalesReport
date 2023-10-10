"""A SIMPLE CLI PROGRAM TO AUTOMATE MONTHLY SALES REPORT
ENCODING IN EXCEL. THIS PROGRAM HAS TWO OPTIONS,
CREATE A NEW REPORT AND EDIT OLD REPORT.

# CREATE A NEW REPORT;
# WILL RUN A LOOP OF INPUT TO GET EVERY ITEM SOLD
# ALONG WITH QUANTITY AND PRICE
# SHOW INPUTED ITEM BEFORE SAVING OR WRITING IT TO EXCEL
# PROGRAM HAS ALSO A FEATURED OF BACK TO REINPUT THE WRONG ITEM

# EDIT OLD REPORT;
# SHOW ALL EXCEL FILE IN THE CURRENT DIRECTORY
# SELECT EXCEL REPORT TO EDIT
# RUN A LOOP OF INPUT TO GET EVERY ITEM SOLD
# ALONG WITH QUANTITY AND PRICE
# SHOW INPUTED ITEM BEFORE SAVING OR WRITING IT TO EXCEL
# PROGRAM HAS ALSO A FEATURED OF BACK TO REINPUT THE WRONG ITEM

-G0DFR3YP4DU4"""

import os
import shutil
from datetime import datetime
from typing import List
import openpyxl
from openpyxl import Workbook, load_workbook


# ALL CODE CATEGORY
CODE_CATEGORIES = [
    "ACC", "AVR", "CAB", "CAM", "CASE", "CEL",
    "CHAIR", "CI", "CONSOLE", "CPU", "DESKTOP", "DVDRW",
    "F", "FAN", "FREE", "GC", "GREASE", "HD",
    "HDEX", "HS", "INK", "KB", "LT", "MB",
    "MEM", "MIC", "MN", "MOUSE", "MOUSE PAD", "NET",
    "OS", "PENTABLET", "PRJCTR", "PRNTR", "PROMO", "PS",
    "PWRSTRIP", "RAM", "SCNR", "SERVER", "SI", "SOFT",
    "SP", "SRVC", "SSD", "SSDEX", "SURVEILLANCE", "TABLET",
    "UD", "UPS", "VC", "VR"]

# ALL LAPTOP BRAND
LT_BRANDS = [
    "ACER", "ASUS", "DELL", "GIGABYTE",
    "HP", "HUAWEI", "INTEL", "LENOVO", "MSI"]

# ALL SERVICES CODE
SERVICES = [
    "ASSEMBLY-DT BASIC", "ASSEMBLY-DT W/CS", "CLEANING DT",
    "CLEANING LT", "CLEANING PRNT", "CRIMPING FEE",
    "DELIVERY FEE", "DS-DT/LT FEE", "DS-INDIV FEE",
    "DS-PRNTR FEE", "FREIGHT FEE", "INSTL-ANTI VIRUS",
    "INSTL-DRIVERS", "INSTL-DT H-END", "INSTL-DT-ACC",
    "INSTL-LT ACC", "INSTL-LT KB/LCD", "INSTL-MS OFFICE",
    "INSTL-OS", "LOST RECEIPT SI", "LOST RECEIPT WS",
    "MARKUP FEE", "MISC FEE", "ON SITE SERVICE FEE",
    "PRINTER REPAIR FEE", "RE-ASSEMBLY DT", "REPAIR DT",
    "REPAIR LT", "RMA SERVICE FEE", "STORAGE FEE",
    "TESTING FEE"]

# ITEMCODE TO BE PUT ON EXCEL BASE ON USER INPUT
ITEMCODES: List[str] = []


class Main:
    """MAIN LOOP OF FUNCTION TO GET THE DATA AS AN INPUT FROM THE USER"""

    def __init__(self):
        # DETAILS OF PURCHASED PRODUCT
        # ['ITEMCODE', 'QTY', 'AMOUNT']
        self.item = []

        # LIST OF ALL ITEM
        self.items = []

        # LIST OF ALL LAPTOP ITEM
        self.lt_items = []

        # LOOP HANDLER
        self.loop = False

        # SAVE HANDLER
        self.save = False

        # BACK HANDLER
        self.back_handler = {
            "back2create": False,
            "back2date": False,
            "back2itemcode": False,
            "back2qty": False,
            "back2amount": False,
        }

    def get_date(self):
        """GET THE DATE BY VEFIRYING IF A VALID DATE
        THEN RETURN THE DATE;
        IF '-' IS INPUTED MEANS BACK"""

        while True:
            try:
                date = input("\nDate (MM/DD/YYYY): ")
            except (KeyboardInterrupt, UnboundLocalError):
                continue

            # BACK TO CREATE OPTION
            if date == "-":
                self.back_handler["back2create"] = True
                date = None
                break

            try:
                # CHECK DATE FORMAT
                datetime.strptime(date, "%m/%d/%Y")
                break
            except ValueError:
                print("\nINVALID DATE. PLEASE TRY AGAIN!")

        return date

    def back(self):
        """REMOVE AMOUNT IN ITEM, BACK TO GET_AMOUNT"""

        if self.loop:
            try:
                del self.item[-1]
                self.back_handler["back2amount"] = True
                self.loop = False
            except IndexError:
                pass

        # BACK TO GET_DATE()
        else:
            self.back_handler["back2date"] = True
            self.items = []
            self.item = []

    def new_lt(self, itemcode, lt_brand):
        """NEW LAPTOP BRAND DETECTOR,
        ASK USER IF THE NEW BRAND TO ADD OR NOT"""

        while True:
            print("\nNEW LAPTOP BRAND IS DETECTED!")
            try:
                add_lt = input("\nAdd this new laptop brand? (Y/N): ").upper()
            except (KeyboardInterrupt, UnboundLocalError):
                continue

            if add_lt == "Y":
                LT_BRANDS.append(lt_brand)
                self.item.append(itemcode.split()[0])
                add_new_lt = True
                break

            if add_lt == "N":
                print("\nPLEASE PUT THE RIGHT ITEMCODE!")
                add_new_lt = False
                break

            # IF add_lt NOT Y/N; INVALID INOUT
            print("INVALID INPUT. PLEASE TRY AGAIN!")

        return add_new_lt

    def new_itemcode(self, itemcode, code_category):
        """NEW ITEMCODE DETECTOR,
        ASK USER IF NEW BRAND WANTS TO ADD OR NOT"""

        add_itemcode = False

        while True:
            print("\nNEW ITEMCODE IS DETECTED!")
            try:
                add_item = input("\nAdd this new itemcode? (Y/N): ").upper()
            except (KeyboardInterrupt, UnboundLocalError):
                continue

            if add_item == "Y":
                CODE_CATEGORIES.append(code_category)
                self.item.append(itemcode.split()[0])
                add_itemcode = True
                break

            if add_item == "N":
                print("PLEASE PUT THE RIGHT ITEMCODE")
                add_itemcode = False
                break

            # IF add_item NOT Y/N; INVALID INPUT
            print("\nINVALID INPUT. PLEASE TRY AGAIN!")

        return add_itemcode

    def get_itemcode(self):
        """FUNCTION TO GET ITEMCODE;
        1. CHECK ITEMCODE FIRST IF IT IS IN THE LIST OF ALL VALID CODE
        2. IF NOT IN THERE, ADD IT TO THE ALL CODE"""

        # set self.back_handler['back2itemcode'] to False
        self.back_handler["back2itemcode"] = False

        self.back_handler["back2date"] = False

        while True:
            try:
                itemcode = input("\nInput ItemCode: ").upper()

                if len(itemcode) == 0:
                    continue

                # 'LT' in 'LT-ACER...'
                code_category = itemcode.split("-")[0]

            except (IndexError, KeyboardInterrupt, UnboundLocalError):
                print("INVALID INPUT. PLEASE TRY AGAIN!")

            if itemcode == "RESET":
                self.item = []
                self.items = []
                print("\nDATA HAS BEEN RESET")
                continue

            if itemcode == "-":
                # Back to self.get_date() or;
                # Back to self.get_amount
                self.back()
                break

            if itemcode == "SAVE":
                # SAVE TRANSACTION
                self.save = True
                break

            if itemcode in SERVICES or code_category in CODE_CATEGORIES:
                # GET LAPTOP BRAND
                try:
                    lt_brand = itemcode.split()[0].split("-")[1]
                except IndexError:
                    pass

                # NEW LAPTOP BRAND DETECTED
                if itemcode.startswith("LT-") and lt_brand not in LT_BRANDS:
                    # Call self.new_lt() to ask if user
                    # want to add new laptop brand
                    add_new_lt = self.new_lt(itemcode, lt_brand)

                    if not add_new_lt:
                        continue
                    break

                # ADD ITEMCODE TO ITEM
                self.item.append(itemcode.split()[0])
                break

            # ITEMCODE IS NOT IN THE LIS
            add_new_item = self.new_itemcode(itemcode, code_category)
            if not add_new_item:
                continue
            break

    def get_qty(self):
        """GET QUANTITY BY RUNNING LOOP
        VERIFY IF VALID, CONTINUE LOOP IF NOT"""

        # set self.back_handler['back2qty'] to False
        self.back_handler["back2qty"] = False

        while True:
            try:
                qty = input("\nInput Quantity: ")

                # DELETE QTY ADDED IN ITEM;
                # BACK TO ITEMCODE FIELD
                if qty == "-":
                    try:
                        del self.item[-1]
                    except IndexError:
                        pass
                    self.back_handler["back2itemcode"] = True
                    break

                # CONVERT QTY INTO INTEGER
                qty = int(qty)
                if qty <= 0:
                    print("\nINVALID INPUT. PLEASE TRY AGAIN")
                    continue

                self.item.append(qty)
                break

            except (ValueError, KeyboardInterrupt, UnboundLocalError):
                print("\nINVALID INPUT. PLEASE TRY AGAIN")
                continue

    def get_amount(self):
        """GET AMOUNT BY RUNNING LOOP
        VERIFY IF VALID THEN BREAK, CONTIRNUE LOOP IF NOT"""

        # set self.back_handler['back2amount'] to False
        self.back_handler["back2amount"] = False

        while True:
            try:
                amount = input("\nInput Amount: ")
                if amount == "-":
                    try:
                        del self.item[-1]
                    except IndexError:
                        pass
                    self.back_handler["back2qty"] = True
                    break

                # CONVERT QTY INTO INTEGER
                amount = float(amount)
                self.item.append(amount)

                # ACTIVATE LOOP BOOLEAN TO ASK ANOTHER ITEMCODE,
                # QUANTITY AND AMOUNT
                self.loop = True
                break

            except (ValueError, KeyboardInterrupt, UnboundLocalError):
                print("\nINVALID INPUT. PLEASE TRY AGAIN")
                continue

    def to_save(self, data):
        """FUNCTION TO DISPLAY THE INPUTED DATA BEFORE SAVING"""

        # clear the terminal
        os.system("cls")

        # Get the terminal width
        terminal_width = shutil.get_terminal_size().columns

        # Divide the terminal width into three columns
        col_width = terminal_width // 3

        # Add column headers
        headers = ["CODE", "QTY", "AMOUNT"]
        output = " | ".join(
                header.center(col_width)
                for header in headers) + "\n"

        # Get total amount of all data and add on data list
        total_amt = sum(int(sub_data[2]) for sub_data in data)
        total_qty = sum(sub_data[1] for sub_data in data)
        total = ["TOTAL", f"{total_qty}", f"{total_amt:,.2f}"]
        data.append(total)

        # Format amount from '100.0' to '100.00'
        # for better viewing
        for sub_data in data[:-1]:
            sub_data[2] = f"{sub_data[2]:,.2f}"

        # Add data rows
        for row in data:
            output += (
                    " | ".join(
                        str(sub_item).center(col_width)
                        for sub_item in row) + "\n"
                    )

        # Convert amount from string to int type
        for sub_data in data[:-1]:
            sub_data[2] = int(sub_data[2].replace(",", "").split(".")[0])

        # Print the output
        print(output)

    def save_transaction(self, date):
        """CONFIRM THE DATA TO BE SAVED BY DISPLAYING ALL THE
        DATA INPUTTED, ASK USER IF WANT TO SAVE OR NOT"""

        while True:
            data = [
                sub_item
                for sub_item in self.items
                if sub_item and len(sub_item) == 3
                ]

            # PRINT DATA TO BE SAVED
            self.to_save(data)

            # FIX DATA TO BE SAVED
            self.lt_items = [
                sub_data
                for sub_data in data
                if sub_data[0].startswith("LT-")
                ]

            self.items = [
                sub_d
                for sub_d in data
                if len(sub_d) == 3 and sub_d[0] != "TOTAL"
                ]

            try:
                add = input(f"Save Transaction for {date}? (Y/N): ").upper()
            except (KeyboardInterrupt, UnboundLocalError):
                pass

            if add == "Y":
                self.save = True
                break

            if add == "N":
                self.save = False
                break

    def continue_loop(self):
        """A SMALL FUNCTION TO APPEND
        CONTINUE LOOP OR CONTINUE GETTING ITEM LIST"""

        self.items.append(list(self.item[:3]))
        try:
            del self.item[:3]
        except IndexError:
            pass
        self.loop = False

    def main(self):
        """MAIN FUNCTION FOR GETTING THE 'date',
        'itemcode', 'quantity' and 'amount' THRU LOOP"""

        print('\n"save" to save transaction')
        print('"reset" to remove all inputed data')
        print('"-" to back into previous text field')

        while True:
            date = self.get_date()

            # Main loop for getting
            # self.item = [itemcode, quantity, amount]
            while not self.back_handler["back2create"]:
                # Get the itemcode
                self.get_itemcode()

                # If user want to back in get_date()
                if self.back_handler["back2date"]:
                    break

                # If self.loop boolean has been activated or True;
                # Append self.item list to self.items
                # Reset self.item list
                if self.loop and not self.back_handler["back2amount"]:
                    self.continue_loop()

                # Call self.save_transaction()
                # if user want to save all entry
                if self.save:
                    self.save_transaction(date)
                    if self.save:
                        break
                    continue

                # Get the quantity
                while not self.loop:
                    if not self.back_handler["back2amount"]:
                        self.get_qty()

                        # If user want to back in self.get_itemcode()
                        if self.back_handler["back2itemcode"]:
                            break

                    # Get the amount
                    while not self.loop:
                        self.get_amount()

                        # If user want to back in self.get_qty()
                        if self.back_handler["back2qty"] or self.loop:
                            break

            if self.save or self.back_handler["back2create"]:
                break

        return self.save, date, self.items, self.lt_items


def write_oldcode(cell, worksheet, date_row, qty, amount):
    """A SIMPLE FUNCTION TO WRITE DATA ON EXCEL
    WHERE ITEMCODE IS ALREADY THERE."""
    qty_col = cell.column
    amount_col = cell.column + 1

    prev_qty = worksheet.cell(column=qty_col, row=date_row).value
    if prev_qty is not None:
        qty = int(prev_qty) + int(qty)
    worksheet.cell(column=qty_col, row=date_row, value=qty)

    prev_amount = worksheet.cell(column=amount_col, row=date_row).value
    if prev_amount is not None:
        amount = int(prev_amount) + int(amount)
    worksheet.cell(column=amount_col, row=date_row, value=amount)

    return worksheet


def update_values(worksheet, row_info):
    """UPDATE VALUES IN EXCEL"""

    qty_col = row_info["qty_col"]
    amt_col = row_info["amt_col"]
    date_row = row_info["date_row"]
    qty = row_info["qty"]
    amount = row_info["amount"]

    prev_qty = worksheet.cell(column=qty_col, row=date_row).value
    if prev_qty is not None:
        qty = int(prev_qty) + int(qty)
    worksheet.cell(column=qty_col, row=date_row, value=qty)

    prev_amount = worksheet.cell(column=amt_col, row=date_row).value
    if prev_amount is not None:
        amount = int(prev_amount) + int(amount)
    worksheet.cell(column=amt_col, row=date_row, value=amount)

    return worksheet


def write_newcode(**kwargs):
    """A SIMPLE FUNCTION TO WRITE NEW DATA OR ITEMCODE IN EXCEL"""
    code = kwargs["code"]
    itemcode = kwargs["itemcode"]
    worksheet = kwargs["worksheet"]
    date_row = kwargs["date_row"]
    qty = kwargs["qty"]
    amount = kwargs["amount"]

    itemcode.append(code)
    itemcode = sorted(itemcode)

    # get the index of the code in itemcode
    code_index = itemcode.index(code)

    # get the worksheet column where the code is going to be insert
    code_column = code_index * 2 + 2
    qty_col = code_column
    amt_col = code_column + 1

    # insert the new itemcode
    worksheet.insert_cols(code_column, 2)
    header = worksheet.cell(row=1, column=code_column, value=code)
    header.alignment = openpyxl.styles.Alignment(horizontal="center")
    worksheet.merge_cells(
        start_row=1,
        start_column=code_column,
        end_row=1,
        end_column=code_column + 1
        )

    qty_header = worksheet.cell(row=2, column=qty_col, value="Qty")
    qty_header.alignment = openpyxl.styles.Alignment(horizontal="center")
    amount_header = worksheet.cell(row=2, column=amt_col, value="Amount")
    amount_header.alignment = openpyxl.styles.Alignment(horizontal="center")

    # add quantity and amount on the same column of data and same row date
    row_info = {
        "qty_col": qty_col,
        "amt_col": amt_col,
        "date_row": date_row,
        "qty": qty,
        "amount": amount,
    }

    update_values(worksheet, row_info)

    return worksheet


def write(worksheet, date, data):
    """FUNCTION TO WRITE DATA IN EXCEL IN All_Items worksheet"""

    # get all the itemcode present in all_items worksheet
    if worksheet.title == "All Items":
        itemcode = [cell.value for cell in worksheet[1] if cell.value]
        itemcode = itemcode[1:-1]

        # fix data list
        data = [[i[0].split("-")[0]] + i[1:] for i in data]

    elif worksheet.title == "Laptop Items":
        itemcode = [cell.value for cell in worksheet[1] if cell.value]
        itemcode = itemcode[1:-1]

        # fix data list
        data = [
                [i[0].split("-")[1]] + i[1:]
                for i in data
                if i[0].startswith("LT-")
                ]

    # add date
    date_found = False
    for cell in worksheet["A"]:
        if cell.value == date:
            date_row = cell.row
            date_found = True
            break

    # if date is not in date column
    if not date_found:
        date_row = worksheet.max_row + 1

    worksheet.cell(row=date_row, column=1, value=date)

    # add data on the same row of date
    for sub_data in data:
        code = sub_data[0]
        qty = sub_data[1]
        amount = sub_data[2]

        for cell in worksheet[1]:
            # write data if code is already in worksheet
            if cell.value == code:
                write_oldcode(cell, worksheet, date_row, qty, amount)

            # write data if code is not in worksheet
            if code not in itemcode:
                write_newcode(
                    code=code,
                    itemcode=itemcode,
                    worksheet=worksheet,
                    date_row=date_row,
                    qty=qty,
                    amount=amount,
                )

    # compute all the qty and amount
    compute(worksheet)

    return worksheet


def compute(worksheet):
    """This function compute the total sales for everyday
    and compute the grand total for every transaction;
        1. Get the worksheet from the parameter
        2. Get the total quantity and amount column
        3. Sum up all the quantity of every item for each day
        4. Sum up all the amount of every item for each day
        5. Write total the total quantity and amount in total sales column"""

    total_qty_col = None
    total_amt_col = None

    # GET TOTAL QTY AND AMOUNT COLUMN
    for cell in worksheet[1]:
        if cell.value and cell.value.upper() == "DAILY SALES":
            total_qty_col = cell.column
            total_amt_col = cell.column + 1

    for row_idx, _ in enumerate(worksheet.iter_rows(min_row=3), start=3):
        # GET THE TOTAL QTY OF CURRENT DATE
        total_qty = sum(
            cell.value
            for col in worksheet.iter_cols(
                min_col=2,
                max_col=worksheet.max_column - 2,
                min_row=row_idx,
                max_row=row_idx,
            )
            for cell in col
            if cell.column % 2 == 0
            if cell.value
        )

        # GET THE TOTAL AMOUNT OF CURRENT DATE
        total_amount = sum(
            cell.value
            for col in worksheet.iter_cols(
                min_col=2,
                max_col=worksheet.max_column - 2,
                min_row=row_idx,
                max_row=row_idx,
            )
            for cell in col
            if not cell.column % 2 == 0
            if cell.value
        )

        worksheet.cell(row=row_idx, column=total_qty_col, value=total_qty)
        worksheet.cell(row=row_idx, column=total_amt_col, value=total_amount)

    return worksheet


def create(worksheet, items):
    """TEMPLATE FOR CREATING A REPORT
    1. Create a Date column
    2. Create a column header name from the itemcode
    in conso with sub header of qty and amount
    3. Add total sales column
    3. Save the template"""

    # Add the column headers
    worksheet["A1"] = "Date"
    worksheet.merge_cells("A1:A2")
    worksheet["A1"].alignment = openpyxl.styles.Alignment(
        horizontal="center", vertical="center"
    )

    for i, item in enumerate(sorted(items)):
        col = i * 2 + 2  # Calculate the column index for each item
        header = worksheet.cell(row=1, column=col, value=item)
        header.alignment = openpyxl.styles.Alignment(horizontal="center")
        worksheet.merge_cells(
            start_row=1, start_column=col, end_row=1, end_column=col + 1
        )

        qty_header = worksheet.cell(row=2, column=col, value="Qty")

        qty_header.alignment = openpyxl.styles.Alignment(horizontal="center")
        amount_header = worksheet.cell(row=2, column=col + 1, value="Amount")
        amount_header.alignment = openpyxl.styles.Alignment(
                horizontal="center")

    # Add the "Total Sales" column
    last_col = len(items) * 2 + 1
    total_sales_header = worksheet.cell(
            row=1,
            column=last_col + 1,
            value="DAILY SALES")

    total_sales_header.alignment = openpyxl.styles.Alignment(
            horizontal="center")

    # Add the column headers for the Total Qty and Total Amount sub-columns
    qty_header = worksheet.cell(row=2, column=last_col + 1, value="TOTAL QTY")
    qty_header.alignment = openpyxl.styles.Alignment(horizontal="center")
    amount_header = worksheet.cell(
            row=2,
            column=last_col + 2,
            value="TOTAL AMOUNT"
            )

    amount_header.alignment = openpyxl.styles.Alignment(horizontal="center")

    # Merge the "Total Sales" header across
    # the two sub-columns and center it vertically
    worksheet.merge_cells(
        start_row=1,
        start_column=last_col + 1,
        end_row=1,
        end_column=last_col + 2
    )

    total_sales_header.alignment = openpyxl.styles.Alignment(
        horizontal="center", vertical="center"
    )

    return worksheet


def banner():
    """SMALL FUNCTION TO SHOW BANNER"""
    os.system("cls")
    os.system("clear")

    # Define the text to be printed
    text = "MONTHLY SALES REPORT"
    author = "\t\t\tBy: b4tug4n"

    # Get the size of the terminal
    terminal_size = shutil.get_terminal_size()

    # Calculate the center position for the text
    center_position = (terminal_size.columns - len(text)) // 2

    # Print the text centered on the terminal
    print("\n\n")
    print("\033[1m\033[14m" + " " * center_position + text)
    print(" " * center_position + author)
    print("\n\n")


def get_file():
    """Menu option to create a new template and monthly sales report"""

    # get all the excel file in the current folder to check
    folder = os.getcwd()
    files = os.listdir(folder)
    xl_files = []
    for file in files:
        if file.endswith(".xlsx"):
            name, ext = os.path.splitext(file)
            uppercase_name = name.upper()
            file = uppercase_name + ext
            xl_files.append(file)

    banner()
    print("\n0. Back")
    print("1. Exit")

    file = False
    while True:
        # get the file name
        try:
            file_name = input("\nSave new file as: ")
        except KeyboardInterrupt:
            pass

        if not file_name:
            continue

        if file_name == "0":
            # break the loop and
            # return 0 to back to main menu
            break

        if file_name == "1":
            # break the loop and
            # return 1 to exit the program
            break

        if file_name.upper() + ".xlsx" in xl_files:
            file = file_name.upper() + ".xlsx"
            print(f"\n\n{file} IS ALREADY EXIST. PLEASE CREATE A NEW ONE!")
            continue
        if file_name.upper() + ".xlsx" not in xl_files:
            file_name = file_name.upper() + ".xlsx"
            break

    return file_name


def create_option1():
    """A SIMPLE FUNCTION TO SHOW CREATE INFORMATION"""
    while True:
        file = get_file()
        if file in ("0", "1"):
            break

        main_ = Main()
        save, date, all_data, lt_data = main_.main()

        if save:
            workbook = Workbook()
            worksheet1 = workbook.active
            worksheet1.title = "All Items"

            worksheet2 = workbook.create_sheet(title="Laptop Items")

            create(worksheet1, CODE_CATEGORIES)
            create(worksheet2, LT_BRANDS)

            write(worksheet1, date, all_data)
            write(worksheet2, date, lt_data)

            workbook.save(file)
            print("\nFILE HAS BEEN SAVED SUCCESSFULLY!")
            break

        if not save:
            continue

    return file


def edit_option2():
    """A SIMPLE FUNCTION TO SHOW EDIT OPTION"""
    while True:
        # display all excel file
        folder = os.getcwd()
        files = os.listdir(folder)

        xl_files = []
        for file in files:
            if file.endswith(".xlsx"):
                xl_files.append(file)
        print("\n")
        print("0. Back")
        for exel in xl_files:
            print(str(xl_files.index(exel) + 1) + ".", exel)

        # select file to edit
        while True:
            try:
                xl_file = int(input("\nSelect file to edit: "))

                if xl_file > len(xl_files) or xl_file < 0:
                    print("\nINVALID OPTION PLEASE TRY AGAIN!")
                    continue
                break
            except KeyboardInterrupt:
                pass
            except ValueError:
                print("\nINVALID OPTION PLEASE TRY AGAIN!")
                continue

        if xl_file == 0:
            file = 0
            break

        file = xl_files[xl_file - 1]

        main_ = Main()
        save, date, all_data, all_lt = main_.main()
        if save:
            workbook = load_workbook(file)
            worksheet1 = workbook["All Items"]
            worksheet2 = workbook["Laptop Items"]

            write(worksheet1, date, all_data)
            write(worksheet2, date, all_lt)

            workbook.save(file)
            print("\nFILE HAS BEEN SAVED SUCCESSFULLY!")
            break
        if not save:
            continue

    return file


def main_menu():
    """MENU FUNCTION FOR SELECTING WHAT TO DO"""
    banner()
    while True:
        print("\n1. Create a new report")
        print("2. Edit old report")
        print("3. Exit")
        try:
            option = input("\nChoose option: ")
        except KeyboardInterrupt:
            pass
        if option == "1":
            opt = create_option1()
            if opt == "0":
                continue
            if opt == "1":
                break
            break
        if option == "2":
            edit_op = edit_option2()
            if edit_op == 0:
                continue
            break
        if option == "3":
            break

        # if option is not 1, 2 or 3; invalid
        print("\nINVALID OPTION PLEASE TRY AGAIN!")


if __name__ == "__main__":
    main_menu()
